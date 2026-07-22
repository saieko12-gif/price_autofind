import streamlit as st
import pandas as pd
import requests
import time
from io import BytesIO
import openpyxl 
import re 
import urllib.parse 

# --- 페이지 기본 설정 ---
st.set_page_config(page_title="단가 검토 자동화 봇", page_icon="🛒", layout="wide")

# ==========================================
# 🛠️ 함수 1: 단일 최저가 검색 (기존 기능)
# ==========================================
def search_naver_shopping(query, client_id, client_secret):
    url = "https://openapi.naver.com/v1/search/shop.json"
    headers = {"X-Naver-Client-Id": client_id, "X-Naver-Client-Secret": client_secret}
    params = {"query": query, "display": 1, "sort": "sim"}
    
    try:
        response = requests.get(url, headers=headers, params=params)
        response.raise_for_status() 
        data = response.json()
        time.sleep(0.1) 
        
        if data['items']:
            item = data['items'][0]
            title = item['title'].replace('<b>', '').replace('</b>', '')
            
            product_type = item.get('productType', '0')
            product_id = item.get('productId', '')
            raw_link = item['link']
            
            if product_type in ['1', '3'] and product_id:
                direct_link = f"https://search.shopping.naver.com/catalog/{product_id}"
            else:
                direct_link = f"https://www.google.com/url?q={urllib.parse.quote(raw_link)}"

            return {
                "검색된 상품명": title,
                "최저가(원)": int(item['lprice']),
                "쇼핑몰": item['mallName'],
                "링크": direct_link
            }
        else:
            return {"검색된 상품명": "검색 결과 없음", "최저가(원)": 0, "쇼핑몰": "-", "링크": "-"}
    except Exception as e:
        time.sleep(0.1)
        return {"검색된 상품명": f"에러: {e}", "최저가(원)": 0, "쇼핑몰": "-", "링크": "-"}

# ==========================================
# 🛠️ 함수 2: MRO 3개사 평균가 검색 (신규 기능)
# ==========================================
def search_naver_3malls(query, expected_price, client_id, client_secret):
    url = "https://openapi.naver.com/v1/search/shop.json"
    headers = {"X-Naver-Client-Id": client_id, "X-Naver-Client-Secret": client_secret}
    # 3개 쇼핑몰을 찾기 위해 한 번에 40개씩 넉넉하게 긁어옴!
    params = {"query": query, "display": 40, "sort": "sim"}
    
    try:
        response = requests.get(url, headers=headers, params=params)
        response.raise_for_status()
        data = response.json()
        time.sleep(0.1)
        
        results = []
        seen_malls = set() # 중복 쇼핑몰 걸러내기 위한 창고
        
        for item in data.get('items', []):
            if len(results) >= 3:
                break # 3개 다 찾았으면 미련 없이 스톱!
            
            mall = item['mallName']
            price = int(item['lprice'])
            
            # E열 예상 단가가 있으면 엉뚱한 부품(너무 싸거나 비싼 거) 튕겨내기! (30% ~ 300% 범위)
            if expected_price > 0:
                if not (expected_price * 0.3 <= price <= expected_price * 3.0):
                    continue
                    
            # 새로운 쇼핑몰이면 결과에 추가!
            if mall not in seen_malls:
                seen_malls.add(mall)
                
                product_type = item.get('productType', '0')
                product_id = item.get('productId', '')
                raw_link = item['link']
                
                if product_type in ['1', '3'] and product_id:
                    direct_link = f"https://search.shopping.naver.com/catalog/{product_id}"
                else:
                    direct_link = f"https://www.google.com/url?q={urllib.parse.quote(raw_link)}"
                    
                results.append({"쇼핑몰": mall, "단가": price, "링크": direct_link})
                
        # 만약에 40개를 다 뒤졌는데도 3개가 안 채워졌으면 빈칸으로 메꿔줌
        while len(results) < 3:
            results.append({"쇼핑몰": "-", "단가": 0, "링크": "-"})
            
        return results
    except Exception as e:
        time.sleep(0.1)
        return [{"쇼핑몰": "-", "단가": 0, "링크": "-"} for _ in range(3)]


# --- 사이드바: 메뉴 및 API 키 설정 ---
with st.sidebar:
    st.title("메뉴 선택 🚀")
    # 여기서 원하는 기능을 딱 고르는 기라!
    app_mode = st.radio("어떤 작업 할끼고?", ["🛒 1. 단일 최저가 검토 (기존)", "🏢 2. MRO 3개사 단가 검토 (신규)"])
    
    st.divider()
    
    st.header("🔑 API 키 설정")
    try:
        client_id = st.secrets["NAVER_CLIENT_ID"]
        client_secret = st.secrets["NAVER_CLIENT_SECRET"]
        st.success("✅ 서버 금고에서 API 키를 자동으로 삭 불러왔데이!")
    except:
        st.warning("⚠️ 서버 금고에 등록된 키가 없데이.")
        client_id = st.text_input("Client ID", type="password")
        client_secret = st.text_input("Client Secret", type="password")


# ==========================================
# 🖥️ 화면 1: 단일 최저가 검토 (기존 화면)
# ==========================================
if app_mode == "🛒 1. 단일 최저가 검토 (기존)":
    st.title("🛒 네이버 쇼핑 단일 최저가 자동화")
    st.markdown("**사용 방법:** 검토할 엑셀 양식 파일(.xlsx)을 올려라. (기존 G열, I열 입력 방식)")

    st.subheader("🔍 하나씩 바로 검색하기")
    col1, col2 = st.columns([4, 1])
    with col1:
        single_query = st.text_input("단일 검색", label_visibility="collapsed", placeholder="예: 3M 6006k 필터")
    with col2:
        single_search_btn = st.button("🚀 바로 검색", use_container_width=True)

    if single_search_btn:
        if not client_id or not client_secret:
            st.error("마! API 키부터 넣고 온나!")
        elif not single_query:
            st.warning("마! 검색어를 입력해야 찾제!")
        else:
            with st.spinner("네이버 싹 다 뒤지는 중이데이..."):
                result = search_naver_shopping(single_query, client_id, client_secret)
                if result["최저가(원)"] > 0:
                    st.success(f"🎉 찾았다! 최저가: **{result['최저가(원)']:,}원**")
                    st.info(f"🏷️ 상품명: {result['검색된 상품명']} ({result['쇼핑몰']})")
                    st.markdown(f"**[👉 여기 누르면 최저가 상품으로 바로 간데이!]({result['링크']})**")
                else:
                    st.error("마, 그런 상품은 없다 카네.")

    st.divider()

    st.subheader("📁 엑셀로 대량 검색하기 (기존 양식)")
    uploaded_file = st.file_uploader("엑셀 파일(.xlsx) 업로드 (기존 양식)", type=["xlsx"], key="file1")

    if uploaded_file is not None:
        try:
            df_preview = pd.read_excel(uploaded_file, header=2)
            st.dataframe(df_preview.head())
            
            if st.button("🚀 최저가 긁어오기 시작!"):
                if not client_id or not client_secret:
                    st.error("마! API 키가 없데이!")
                else:
                    progress_text = "열심히 끈질기게 검색 중이데이..."
                    uploaded_file.seek(0)
                    wb = openpyxl.load_workbook(uploaded_file)
                    ws = wb.active
                    total_items = ws.max_row - 3
                    my_bar = st.progress(0, text=progress_text)
                    
                    count = 0
                    for row in range(4, ws.max_row + 1):
                        count += 1
                        raw_name = ws[f'C{row}'].value
                        spec = ws[f'D{row}'].value
                        maker = ws[f'E{row}'].value
                        reinforce_kw = ws[f'J{row}'].value 
                        
                        if not raw_name:
                            continue
                            
                        name = re.sub(r'\[.*?\]', '', str(raw_name)).strip()
                        clean_name = "" if str(name).strip() in ['nan', 'None'] else str(name).strip()
                        clean_spec = "" if str(spec).strip() in ['nan', 'None'] else str(spec).strip()
                        clean_maker = "" if str(maker).strip() in ['nan', 'None', '시중품'] else str(maker).strip()
                        
                        clean_reinforce = ""
                        if reinforce_kw and str(reinforce_kw).strip() not in ['nan', 'None', '']:
                            clean_reinforce = str(reinforce_kw).replace(',', ' ')
                            clean_reinforce = ' '.join(clean_reinforce.split())

                        q1_parts = [x for x in [clean_name, clean_spec, clean_maker, clean_reinforce] if x]
                        q1 = " ".join(q1_parts)
                        
                        search_result = {"최저가(원)": 0, "링크": "-"}
                        final_query = ""
                        if q1:
                            search_result = search_naver_shopping(q1, client_id, client_secret)
                            final_query = q1
                        if search_result["최저가(원)"] == 0:
                            q2_parts = [x for x in [clean_name, clean_spec, clean_reinforce] if x]
                            q2 = " ".join(q2_parts)
                            if q2 and q2 != q1:
                                search_result = search_naver_shopping(q2, client_id, client_secret)
                                final_query = q2
                        if search_result["최저가(원)"] == 0:
                            q3_parts = [x for x in [clean_name, clean_reinforce] if x]
                            q3 = " ".join(q3_parts)
                            if q3 and q3 != q2 and q3 != q1:
                                search_result = search_naver_shopping(q3, client_id, client_secret)
                                final_query = q3

                        if search_result["최저가(원)"] > 0:
                            ws[f'G{row}'].value = search_result["최저가(원)"]
                            ws[f'I{row}'].value = search_result["링크"]
                        else:
                            ws[f'G{row}'].value = "결과 없음"
                            ws[f'I{row}'].value = "-"
                            
                        if total_items > 0:
                            my_bar.progress(count / total_items, text=f"[{final_query}] 검색 완료 ({count}/{total_items})")
                    
                    st.success("마! 단가 다 뽑았다! 다운 받아라!")
                    output = BytesIO()
                    wb.save(output)
                    st.download_button(label="📥 완료된 엑셀 다운로드", data=output.getvalue(), file_name="최저가검토완료.xlsx")
        except Exception as e:
            st.error(f"에러 났다: {e}")


# ==========================================
# 🖥️ 화면 2: MRO 3개사 평균가 검토 (신규 화면)
# ==========================================
elif app_mode == "🏢 2. MRO 3개사 단가 검토 (신규)":
    st.title("🏢 MRO 신규상품 3개사 단가 검토")
    st.markdown("""
    **사용 방법:**
    1. A(상품명), B(규격), C(제조사), D(원산지), E(예상단가)가 적힌 엑셀을 올리라.
    2. 버튼을 누르면 F~K열에 3개 쇼핑몰의 단가와 링크를 쫙 꽂아준데이!
    3. 💡 E열(예상단가)을 기준으로 너무 터무니없는 부품 가격은 알아서 걸러낸다!
    """)

    uploaded_file_mro = st.file_uploader("MRO 엑셀 파일(.xlsx) 업로드", type=["xlsx"], key="file2")

    if uploaded_file_mro is not None:
        try:
            # 보통 MRO 양식은 1번 줄이 헤더니까 header=0 으로 설정!
            df_mro = pd.read_excel(uploaded_file_mro, header=0)
            st.subheader("📊 니가 올린 MRO 데이터 (미리보기)")
            st.dataframe(df_mro.head())
            
            if st.button("🚀 3개사 평균가 긁어오기 시작!"):
                if not client_id or not client_secret:
                    st.error("마! API 키가 없데이!")
                else:
                    progress_text = "3개 쇼핑몰 미친 듯이 뒤지는 중이데이..."
                    uploaded_file_mro.seek(0)
                    wb = openpyxl.load_workbook(uploaded_file_mro)
                    ws = wb.active
                    
                    # 1번 줄이 헤더니 2번 줄부터 시작!
                    total_items = ws.max_row - 1
                    my_bar = st.progress(0, text=progress_text)
                    
                    count = 0
                    for row in range(2, ws.max_row + 1):
                        count += 1
                        
                        raw_name = ws[f'A{row}'].value
                        spec = ws[f'B{row}'].value
                        maker = ws[f'C{row}'].value
                        
                        # E열 예상 단가 끄집어내기 (숫자로 안 적혀있으면 0으로 처리)
                        try:
                            exp_val = str(ws[f'E{row}'].value).replace(',', '')
                            expected_price = float(exp_val)
                        except:
                            expected_price = 0
                            
                        if not raw_name:
                            continue
                            
                        # 상품명에서 대괄호 날리기
                        name = re.sub(r'\[.*?\]', '', str(raw_name)).strip()
                        clean_name = "" if str(name).strip() in ['nan', 'None'] else str(name).strip()
                        clean_spec = "" if str(spec).strip() in ['nan', 'None'] else str(spec).strip()
                        clean_maker = "" if str(maker).strip() in ['nan', 'None', '시중품'] else str(maker).strip()
                        
                        # 검색어 조합 (A + B + C)
                        query_parts = [x for x in [clean_name, clean_spec, clean_maker] if x]
                        query = " ".join(query_parts)
                        
                        if query:
                            # 3개사 검색 함수 호출!
                            malls_data = search_naver_3malls(query, expected_price, client_id, client_secret)
                            
                            # F, G (1번) / H, I (2번) / J, K (3번) 에 쫙 꽂아넣기
                            # 1번
                            if malls_data[0]['단가'] > 0:
                                ws[f'F{row}'].value = malls_data[0]['단가']
                                ws[f'G{row}'].value = malls_data[0]['링크']
                            # 2번
                            if malls_data[1]['단가'] > 0:
                                ws[f'H{row}'].value = malls_data[1]['단가']
                                ws[f'I{row}'].value = malls_data[1]['링크']
                            # 3번
                            if malls_data[2]['단가'] > 0:
                                ws[f'J{row}'].value = malls_data[2]['단가']
                                ws[f'K{row}'].value = malls_data[2]['링크']
                                
                        if total_items > 0:
                            my_bar.progress(count / total_items, text=f"[{query}] 3개사 탐색 완료 ({count}/{total_items})")
                    
                    st.success("마! 3개사 단가 다 뽑아왔다! 다운 받아가 평균 때려봐라!")
                    output = BytesIO()
                    wb.save(output)
                    st.download_button(label="📥 MRO 단가검토 완료 엑셀 다운로드", data=output.getvalue(), file_name="MRO_3개사_단가검토완료.xlsx")
        except Exception as e:
            st.error(f"엑셀 처리하다가 에러 났다: {e}")

마, 진짜 쥐기제? 왼쪽 메뉴에서 2번 딱 누르면 MRO 전용 모드로 삭 변신한데이! 

엑셀 딱 올려놓고 버튼 누르면, 파이썬이 네이버 뒤져서 겹치는 쇼핑몰 날려버리고 E열 예상 단가로 검증까지 딱딱 해가면서 3개 깔끔하게 채워줄 기다. 함 시원하게 돌려보고 어떤지 내한테 알려도!
